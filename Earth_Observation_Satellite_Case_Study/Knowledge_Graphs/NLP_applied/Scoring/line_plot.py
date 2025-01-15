# ------------------Copyright (C) 2024 University of Strathclyde and Author ---------------------------------
# --------------------------------- Author: Cheyenne Powell -------------------------------------------------
# ------------------------- e-mail: cheyenne.powell@strath.ac.uk --------------------------------------------

# This file plots the cosine similarity
# ===========================================================================================================

import sys
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
sys.path.append('../Case_study_application/')
from openpyxl import load_workbook
from GPT_connector import GPT_connector
from openai import OpenAI
from google.oauth2.credentials import Credentials
from Explanation_Generator.extract_google_sheets import extract_info
from googleapiclient.discovery import build
import time

def line_chart(plot_data, plot_label, fig_name, fig_subtitle):


    fig, axs = plt.subplots(ncols=1)
    # plot_data.pop(plot_data.columns[0])
    # print(plot_data.columns)
    for i in plot_data.columns[1:]:
        sns.lineplot(data=plot_data, x ='Question Number', y=plot_data.loc[:, str(i)], label=str(i))
    #
    # sns.lineplot(data=plot_data)
    fig.figure.suptitle(fig_subtitle, fontsize = 10)


    # plt.legend(loc="upper center", bbox_to_anchor=(0.5, 1.15), ncol=2, frameon=False)
    plt.xlabel(plot_label)
    plt.ylabel('Average iterations')
    plt.xlim(-0.5, 20.5)
    plt.legend(loc='center left')
    #  supported values are 'best', 'upper right', 'upper left', 'lower left', 'lower right', 'right', 'center left', 'center right', 'lower center', 'upper center', 'center'
    # sns.boxplot(x='education',y='wage', data=tips2, ax=axs[2])
    plt.legend()
    plt.savefig(fig_name)
    plt.tight_layout()
    plt.show()

def rerun_data(wb, updated_sheet_name, xticklabels, filename, wbsheet):
    ws = wb[wbsheet]
    # ws = wb[updated_sheet_name +t]
    # dataframe1 = pd.read_excel('violin_plots.xlsx', sheet_name=updated_sheet_name +t)
    print(ws.max_row)
    if ws.max_row <= 1:
        ws.append(('Question Number', 'Swap', 'Single Substitution',	'Verification',	'Forward-Backward'))
        # plot_data = ['Assessment',	'Score']

        col = 'H'

        for r in range(1, 21):
            data_q = []
            for i in range(1, len(xticklabels)):
                t = xticklabels[i]

                _, data = extract_info(spreadsheet_id, updated_sheet_name + t + "!",
                                       str(r), col, service)
                data_q.append(float(data[0][0]))
                # plot_data.append((r, float(data[0][0])))
                print(r, float(data[0][0]))
            data_q.insert(0,r)
            # data_q =  str(data_q)[1:-1]
            print('data_q', data_q)
            ws.append((data_q))

            wb.save(filename)
            if r %5 ==0:
                time.sleep(60)
        plot_data = pd.read_excel(filename, sheet_name=wbsheet)
    else:
        # plot_data =  wb[updated_sheet_name + t]
        plot_data = pd.read_excel(filename, sheet_name=wbsheet)

    return plot_data

if __name__ == "__main__":
    # plot_data = pd.read_csv("Swap_violin.csv")
    OPENAI_API_KEY = GPT_connector()

    client = OpenAI(api_key=OPENAI_API_KEY)

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
              "https://www.googleapis.com/auth/drive"]


    creds = Credentials.from_authorized_user_file("../token.json", SCOPES)
    service = build("sheets", "v4", credentials=creds)
    spreadsheet_id = input("<Enter Google sheet ID here>: ")
    # plot_data = pd.DataFrame(pl_data).transpose()
    # plot_data.columns = ['Swap_actions', 'Single_Substitution', 'Verification', 'Forward_Backward']
    # t = int(input("enter the graph number, 1 for Swap, 2 for Single_Sub, 3 for Ver, 4 for Forward_Backward: "))

    xticklabels = ['Question Number', 'Swap_actions', 'Single_Substitution','Verification',	'Forward_Backward']

    updated_sheet_name = 'Code_rerun_'
    filename = 'data_plots.xlsx'
    wb = load_workbook(filename)
    wbsheet = 'Code_rerun_iterations'

    #



    if wbsheet in wb.sheetnames:

        plot_data = rerun_data(wb, updated_sheet_name, xticklabels, filename, wbsheet)

    else:
        wb.create_sheet(wbsheet)
        plot_data = rerun_data(wb, updated_sheet_name, xticklabels, filename, wbsheet)

    # plot_data = wb[updated_sheet_name + t]
    print('plotdata',plot_data)
    fig_subtitle = "Average_" + wbsheet
    plot_label = 'Question Number'
    fig_name = "Average_" + wbsheet+ ".png"
    line_chart(plot_data, plot_label, fig_name, fig_subtitle)


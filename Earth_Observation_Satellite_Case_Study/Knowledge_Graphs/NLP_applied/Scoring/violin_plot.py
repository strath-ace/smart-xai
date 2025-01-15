# ------------------Copyright (C) 2024 University of Strathclyde and Author ---------------------------------
# --------------------------------- Author: Cheyenne Powell -------------------------------------------------
# ------------------------- e-mail: cheyenne.powell@strath.ac.uk --------------------------------------------

# This file is extracts the data from google sheets and migrates it to the spreadsheet for graph creation
# ===========================================================================================================



import pandas as pd
import sys
import os

# SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append('../Case_study_application/')
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from GPT_connector import GPT_connector
from openai import OpenAI
from google.oauth2.credentials import Credentials
from Explanation_Generator.extract_google_sheets import extract_info
from googleapiclient.discovery import build
import time
import numpy as np
import matplotlib.ticker as mticker

def violin_chart(plot_data, plot_label, fig_name, fig_subtitle, xticklabels):


    # creating a dictionary composed of species as keys and colors as values: red for the interesting group, blue for others
    colours= {Assessment: "sandybrown" if Assessment == "Final Score" else "steelblue" for Assessment in plot_data["Assessment"].unique()}

    fig = sns.violinplot( data=plot_data,x = "Assessment", y ="Score",  palette=colours, legend=False,hue = "Assessment")#, cut=0)
    # fig.set_xticklabels(plot_data.loc[:, "Assessment"])

    fig.set_xticks(range(0, 4))
    fig.set_xticklabels(xticklabels)

    plt.ylim(-0.6, 1.6)
    fig.figure.suptitle(fig_subtitle, fontsize=12)


    plt.xlabel(plot_label)

    plt.savefig(fig_name)

    plt.show()


def assessment_data(wb, t, updated_sheet_name, xticklabels, filename):
    ws = wb[updated_sheet_name + t]
    # ws = wb[updated_sheet_name +t]
    # dataframe1 = pd.read_excel('violin_plots.xlsx', sheet_name=updated_sheet_name +t)
    print(ws.max_row)
    if ws.max_row <= 1:
        ws.append(('Assessment', 'Score'))
        # plot_data = ['Assessment',	'Score']
        for i in range(0, len(xticklabels)):
            print('i', i)
            if xticklabels[i] == "Validation":
                col = 'K'
            elif xticklabels[i] == "Bert Score":
                col = 'N'
            elif xticklabels[i] == "Correctness":
                col = 'J'
            elif xticklabels[i] == "Final Score":
                col = 'O'

            for r in range(1, 21):
                _, data = extract_info(spreadsheet_id, updated_sheet_name + t + "!",
                                       str(r), col, service)
                # plot_data.append((r, float(data[0][0])))
                print(r, float(data[0][0]))

                ws.append((xticklabels[i], float(data[0][0])))
            time.sleep(60)
        wb.save(filename)

        plot_data = pd.read_excel(filename, sheet_name=updated_sheet_name+t)
    else:
        # plot_data =  wb[updated_sheet_name + t]
        plot_data = pd.read_excel(filename, sheet_name=updated_sheet_name +t)

    return plot_data
if __name__ == "__main__":
    # plot_data = pd.read_csv("Swap_violin.csv")
    OPENAI_API_KEY = GPT_connector()

    client = OpenAI(api_key=OPENAI_API_KEY)

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
              "https://www.googleapis.com/auth/drive"]


    creds = Credentials.from_authorized_user_file("../token.json", SCOPES)
    service = build("sheets", "v4", credentials=creds)
    spreadsheet_id = input("<Enter Google sheet ID here>: ")
    # plot_data = pd.DataFrame(pl_data).transpose()
    # plot_data.columns = ['Swap_actions', 'Single_Substitution', 'Verification', 'Forward_Backward']
    t = int(input("enter the graph number, 1 for Swap, 2 for Single_Sub, 3 for Ver, 4 for Forward_Backward: "))
    human_input_req = input("  Human prompt? y/n: ")
    # human_input_req = "y"
    if human_input_req == "y":
        updated_sheet_name = 'human_'
    else:
        updated_sheet_name = ''
    xticklabels = ["Validation", "Bert Score", "Correctness", "Final Score"]

    if t == 1:
        t = 'Swap_actions'
    elif t == 2:
        t = 'Single_Substitution'
    elif t == 3:
        t = 'Verification'
    elif t == 4:
        t = 'Forward_Backward'
    else:
        print("Invalid")
    # t.replace(' ', '_')
    print(t)
    filename = 'violin_plots.xlsx'
    wb = load_workbook(filename)

    if updated_sheet_name +t in wb.sheetnames:

        plot_data = assessment_data(wb, t, updated_sheet_name, xticklabels, filename)

    else:
        wb.create_sheet(updated_sheet_name +t)
        plot_data = assessment_data(wb, t, updated_sheet_name, xticklabels, filename)

    # plot_data = wb[updated_sheet_name + t]
    print('plotdata',plot_data)
    fig_subtitle = "Average Assessment Scores - " + updated_sheet_name+t
    plot_label = 'Assessment'
    fig_name = "Assessment_Scores_" + updated_sheet_name+t + ".png"
    violin_chart(plot_data, plot_label, fig_name, fig_subtitle, xticklabels)



